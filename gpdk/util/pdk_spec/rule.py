import shutil
from abc import abstractmethod
from dataclasses import dataclass
from pathlib import Path
from types import ModuleType
from typing import Any, Tuple, Iterable, Union, List

import openpyxl
from openpyxl.styles import Font

from fnpcell.ansi.term import Color, print_ansi, style
from fnpcell.interfaces import IProcessor, IRunnable, IUpdatable
from gpdk.util.pdk_spec.interfaces import ICheckRule, IRuleLib

_T = Union[None, ICheckRule, Iterable["_T"]]


@dataclass
class Progressable(IUpdatable):
    _font: Font = Font(size=12)
    _style: str = style(color=Color.GREEN)
    _title: str = ""
    total: int = 0
    index: int = 0

    @property
    def style(self):
        return self._style

    @style.setter
    def style(self, value: str):
        self._style = value

    @property
    def title(self):
        return self._title

    @title.setter
    def title(self, value: str):
        self._title = value

    def with_total(self, total: int):
        return self.updated(total=total)

    def with_index(self, index: int):
        return self.updated(index=index)

    def progress(self):
        terminal_width = shutil.get_terminal_size().columns
        progress = f"{round((self.index + 1) / self.total * 100)}%"
        progress = f"[{progress.rjust(4)}]".rjust(terminal_width - len(self.title))
        print_ansi(self.title, self.style, progress)


@dataclass
class CheckRule(ICheckRule, Progressable):
    @abstractmethod
    def checked(self, pdk: ModuleType, worksheet: openpyxl.Workbook, index: int) -> Tuple[str, str]:
        pass


@dataclass(frozen=True)
class RuleLib(IRuleLib):
    content: Tuple[CheckRule, ...] = tuple()

    def __add__(self, other: _T) -> "RuleLib":
        content: List[CheckRule] = []
        content.extend(self.content)
        if isinstance(other, RuleLib):
            content.extend(other.content)
        elif other is not None:
            if not isinstance(other, Iterable):
                other = [other]
            for it in other:
                content.append(it)  # type: ignore
        return RuleLib(content=tuple(content))

    def check(self, pdk: ModuleType):
        self.run(processor=RuleChecker(pdk))


class RuleChecker(IProcessor):
    def __init__(self, pdk: ModuleType):
        self.pdk = pdk
        self._terminal_width = shutil.get_terminal_size().columns
        self._workbook = openpyxl.Workbook()
        self._worksheet = self._workbook.active  # type: ignore
        self._worksheet.title = "report"  # type: ignore
        print_ansi(f" Checking {pdk.__name__} ".center(self._terminal_width, "="))

    def enter(self, target: IRunnable) -> bool:
        self._worksheet.cell(1, 1, "检查项目").font = Font(size=16, bold=True, color="00FF6600")  # type: ignore

        if isinstance(target, RuleLib):
            results: List[Tuple[str, str]] = []
            content = target.content
            total = len(content)

            for index, rule in enumerate(content):
                results.append(rule.checked(self.pdk, self._workbook, index + 2))
                new_rule = rule.with_total(total).with_index(index)  # type: ignore
                new_rule.progress()
            flag_results = [r for r in results if r != ("", "")]
            if any(flag_results):
                _print_ansi_result(flag_results, self._workbook, total)
            else:
                print_ansi(style(color=Color.GREEN), " PASSED ".center(self._terminal_width, "="))

            pdk_path = Path(self.pdk.__path__[0]) / f"check_{self.pdk.__name__}_report"
            output_file = pdk_path.with_suffix(".xlsx")
            self._workbook.save(str(output_file))  # type: ignore

        return True

    def exit(self, target: Any) -> None:
        pass


def _print_ansi_result(result: List[Tuple[str, str]], workbook: openpyxl.Workbook, index: int):
    terminal_width = shutil.get_terminal_size().columns
    fails: List[str] = []
    warnings: List[str] = []
    for color, content in result:
        if color == style(color=Color.RED):
            fails.append(content)
        elif color == style(color=Color.YELLOW):
            warnings.append(content)

    font1 = Font(
        size=16,
        bold=True,
        color="00FF6600",
    )
    font2 = Font(size=12)

    total = index
    worksheet = workbook.active  # type: ignore
    worksheet.cell(index + 5, 1, "检查结果统计").font = font1  # type: ignore
    worksheet.cell(index + 6, 1, "Success").font = font2  # type: ignore
    worksheet.cell(index + 6, 2, f"{total - len(fails)-len(warnings)}").font = font2  # type: ignore
    worksheet.cell(index + 7, 1, "Fail").font = font2  # type: ignore
    worksheet.cell(index + 7, 2, f"{len(fails)}").font = font2  # type: ignore
    worksheet.cell(index + 8, 1, "Warning").font = font2  # type: ignore
    worksheet.cell(index + 8, 2, f"{len(warnings)}").font = font2  # type: ignore
    worksheet.cell(index + 12, 1, "检查明细").font = font1  # type: ignore
    next_i = index + 13

    if warnings:
        worksheet.cell(  # type: ignore
            next_i, 1, "----------------------------------------------------WARNING----------------------------------------------------"
        ).font = Font(  # type: ignore
            size=12, color="003366FF"
        )

        print_ansi(style(color=Color.YELLOW), " WARNING ".center(terminal_width, "="))
        for _, warning in enumerate(warnings):
            print_ansi(warning, style(color=Color.YELLOW))

            warning_ = warning.split("\n")
            next_i += 1
            for _, per_warning in enumerate(warning_):
                worksheet.cell(next_i, 1, per_warning).font = font2  # type: ignore
                next_i += 1

    if fails:
        worksheet.cell(next_i, 1, "----------------------------------------------------FAILED----------------------------------------------------").font = Font(  # type: ignore
            size=12, color="00FF0000"
        )

        print_ansi(style(color=Color.RED), " FAILED ".center(terminal_width, "="))
        for _, fail in enumerate(fails):
            print_ansi(fail, style(color=Color.RED))

            fail_ = fail.split("\n")
            next_i += 1
            for _, per_fail in enumerate(fail_):
                worksheet.cell(next_i, 1, per_fail).font = font2  # type: ignore
                next_i += 1
